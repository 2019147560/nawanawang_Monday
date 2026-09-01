"""
충북청년희망센터(cbhope1539.net) 공지사항 스크래퍼
- 은둔고립청년 / 청년일자리 / 청년모임 관련 게시물만 수집
- 전체 공지사항(186건 / 19페이지) 순회 후 키워드 필터링
사용법: python cbhope_scraper.py
결과물: cbhope_events_YYYYMMDD_HHMMSS.json / .xlsx
"""

import re
import time
import json
import random
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

BASE_URL = "https://www.cbhope1539.net"
LIST_URL = "https://www.cbhope1539.net/bbs/list.do"
VIEW_URL = "https://www.cbhope1539.net/bbs/view.do"

LIST_PARAMS = {
    "key":  "2004215717214",
    "gbn":  "noice",
    "sc":   "",
    "sw":   "",
    "orderBy": "",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": BASE_URL,
}

ORG_NAME = "충북청년희망센터"

# ── 관련 키워드 ───────────────────────────────────────────────────────────────
RELEVANT_KEYWORDS = [
    "은둔", "고립", "히키코모리", "사회적고립",
    "청년일자리", "일자리지원", "취업지원", "취업연계", "일경험",
    "청년모임", "자조모임", "커뮤니티", "동아리", "소모임",
    "청년취업", "청년고용", "미취업", "구직",
    "니트", "NEET", "가족돌봄", "고립·은둔", "고립은둔",
]


# ── 유틸 ─────────────────────────────────────────────────────────────────────

def fetch(url: str, params: dict = None) -> BeautifulSoup | None:
    try:
        r = requests.get(url, params=params, headers=HEADERS, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [오류] {url} → {e}")
        return None


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def is_relevant(text: str) -> bool:
    t = text.lower()
    return any(kw.lower() in t for kw in RELEVANT_KEYWORDS)


# ── 1단계: 목록 페이지에서 bbsSn 수집 ────────────────────────────────────────

def get_post_links() -> list[dict]:
    """
    cbhope1539.net 목록 구조:
    - 제목 링크: href="#none" (JS 처리)
    - bbsSn은 tr의 onclick 또는 숨겨진 input에서 추출
    - 또는 a 태그의 data 속성
    전략: 목록 HTML에서 bbsSn 추출 후 view.do URL 직접 생성
    """
    links = []
    seen  = set()

    for page in range(1, 25):  # 최대 25페이지 (현재 19페이지)
        params = {**LIST_PARAMS, "pageIndex": page}
        print(f"  목록 페이지 {page} 파싱 중...")
        soup = fetch(LIST_URL, params=params)
        if not soup:
            break

        # bbsSn 추출 전략들
        found_any = False

        # 전략 1: tr/td 의 onclick에서 bbsSn 추출
        # 예: onclick="goView('521')" 또는 fn_view(521)
        for tag in soup.find_all(attrs={"onclick": True}):
            onclick = tag.get("onclick", "")
            m = re.search(r"[\w_]*[Vv]iew\D*(\d{3,})", onclick)
            if not m:
                m = re.search(r"bbsSn[=,'\s]+(\d{3,})", onclick)
            if not m:
                continue

            bbs_sn = m.group(1)
            if bbs_sn in seen:
                continue

            # 제목 추출 (같은 행의 텍스트)
            row = tag.find_parent("tr") or tag.find_parent("li") or tag
            title = ""
            if row:
                # 제목 셀 (td.subject, td:nth-child(3) 등)
                title_td = row.select_one("td.subject, td.title, .td_subject")
                if title_td:
                    title = clean(title_td.get_text())
                else:
                    title = clean(row.get_text())[:80]

            # 등록일
            date_td = row.select_one("td:last-child, .td_date") if row else None
            pub_date = clean(date_td.get_text()) if date_td else ""

            # 분류
            cat_td = row.select_one("td.cate, td:nth-child(2)") if row else None
            category = clean(cat_td.get_text()) if cat_td else ""

            full_url = (
                f"{VIEW_URL}?bbsSn={bbs_sn}"
                f"&key={LIST_PARAMS['key']}&gbn={LIST_PARAMS['gbn']}&pageIndex={page}"
            )
            seen.add(bbs_sn)
            links.append({
                "url":      full_url,
                "bbs_sn":   bbs_sn,
                "title":    title,
                "pub_date": pub_date,
                "category": category,
                "cover":    "",
            })
            found_any = True

        # 전략 2: 숨겨진 input[name=bbsSn]
        if not found_any:
            for inp in soup.find_all("input", {"name": re.compile(r"bbsSn|no", re.I)}):
                bbs_sn = inp.get("value", "")
                if not bbs_sn or bbs_sn in seen:
                    continue
                row = inp.find_parent("tr") or inp.find_parent("li")
                title = clean(row.get_text())[:80] if row else ""
                full_url = (
                    f"{VIEW_URL}?bbsSn={bbs_sn}"
                    f"&key={LIST_PARAMS['key']}&gbn={LIST_PARAMS['gbn']}&pageIndex={page}"
                )
                seen.add(bbs_sn)
                links.append({"url": full_url, "bbs_sn": bbs_sn, "title": title,
                               "pub_date": "", "category": "", "cover": ""})
                found_any = True

        # 전략 3: a[href*='bbsSn'] 직접 파싱
        if not found_any:
            for a in soup.select("a[href*='bbsSn'], a[href*='view.do']"):
                href = a.get("href", "")
                m = re.search(r"bbsSn=(\d+)", href)
                if not m:
                    continue
                bbs_sn = m.group(1)
                if bbs_sn in seen:
                    continue
                full_url = urljoin(BASE_URL, href)
                seen.add(bbs_sn)
                links.append({"url": full_url, "bbs_sn": bbs_sn,
                               "title": clean(a.get_text()), "pub_date": "",
                               "category": "", "cover": ""})
                found_any = True

        if not found_any:
            print(f"  → 페이지 {page}에 게시물 없음, 중단")
            break

        # 마지막 페이지 확인
        pager = soup.select_one(".paging, .pagination, .page_wrap")
        if pager:
            next_btn = pager.select_one(f"a[href*='pageIndex={page+1}']")
            if not next_btn and page > 1:
                print(f"  → 마지막 페이지 도달 ({page})")
                break

        time.sleep(random.uniform(0.5, 1.0))

    print(f"  → 총 {len(links)}개 게시물 수집")
    return links


# ── 2단계: 상세 페이지 파싱 + 관련성 필터 ────────────────────────────────────

def parse_detail(hint: dict) -> dict | None:
    url  = hint["url"]
    soup = fetch(url)
    if not soup:
        return None

    # ── 상세 테이블 파싱: th/td 구조
    view_table = soup.select_one("table.view_table, .board_view, table")
    rows_map: dict[str, str] = {}
    if view_table:
        for tr in view_table.select("tr"):
            th = tr.select_one("th")
            td = tr.select_one("td")
            if th and td:
                rows_map[clean(th.get_text())] = clean(td.get_text())

    # ── title
    title = (
        rows_map.get("제목") or
        clean(soup.select_one(".view_title, h3.title, h2")
              .get_text()) if soup.select_one(".view_title, h3.title, h2") else ""
    ) or hint.get("title", "")

    # ── 등록일
    pub_date = (
        rows_map.get("등록일") or
        hint.get("pub_date", "")
    )
    # YYYY-MM-DD만 추출
    dm = re.search(r"(20\d{2}-\d{2}-\d{2})", pub_date)
    pub_date = dm.group(1) if dm else pub_date

    # ── 분류
    category = rows_map.get("분류") or hint.get("category", "")

    # ── 본문 (이미지 위주일 수 있으므로 alt + 텍스트 모두 활용)
    content = soup.select_one(".board_view_content, .view_content, .content, td.content")
    content_text = ""
    if content:
        # img alt 텍스트도 포함
        for img in content.select("img[alt]"):
            alt = img.get("alt", "")
            if len(alt) > 3:
                content_text += " " + alt
        content_text = clean(content.get_text(" ")) + content_text

    # ── 관련성 필터
    if not is_relevant(title) and not is_relevant(content_text[:500]):
        return None

    # ── org
    org = ORG_NAME
    m = re.search(r"(주관|주최|운영기관)[:\s]*([^\n,/]{2,30})", content_text)
    if m:
        org = clean(m.group(2))

    # ── region
    region = "충북"
    region_list = ["전국", "청주", "충주", "제천", "보은", "옥천", "영동",
                   "증평", "진천", "괴산", "음성", "단양"]
    for r in region_list:
        if r in content_text:
            region = r
            break

    # ── mode
    if re.search(r"온·오프|온\s*오프|혼합", content_text):
        mode = "혼합"
    elif re.search(r"온라인|비대면|zoom|화상|유튜브", content_text, re.I):
        mode = "온라인"
    elif re.search(r"오프라인|현장|대면|방문|센터", content_text):
        mode = "오프라인"
    else:
        mode = ""

    # ── motive
    motive: list[str] = []
    m = re.search(
        r"(지원\s*대상|신청\s*대상|모집\s*대상|참여\s*대상)[^\n:❍□○]*[:\s❍□○]+([\s\S]{5,300}?)(?=\n\n|❍|□|○|▶|\Z)",
        content_text
    )
    if m:
        items = re.split(r"[\n,◦]", m.group(2))
        motive = [clean(i) for i in items if 3 < len(clean(i)) < 100][:5]

    # ── deadline
    deadline = ""
    m = re.search(
        r"(신청\s*기간|모집\s*기간|접수\s*기간)[^\n:❍□○]*[:\s❍□○]+([\d.\-년월일\s~]+)",
        content_text
    )
    if m:
        deadline = clean(m.group(2))

    # ── status
    if re.search(r"마감|종료|접수종료|모집완료", title + content_text[:300]):
        status = "마감"
    elif re.search(r"예정|추후공지|미정", content_text[:300]):
        status = "모집 예정"
    else:
        now = datetime.now()
        date_m = re.search(r"(20\d{2})[.\-년]\s*(\d{1,2})[.\-월]\s*(\d{1,2})", deadline)
        if date_m:
            try:
                dl = datetime(int(date_m.group(1)), int(date_m.group(2)), int(date_m.group(3)))
                status = "마감" if now > dl else "현재 신청 가능"
            except Exception:
                status = "현재 신청 가능"
        else:
            status = "현재 신청 가능"

    # ── period
    period = ""
    m = re.search(
        r"(사업\s*기간|운영\s*기간|지원\s*기간|프로그램\s*기간)[^\n:❍□○]*[:\s❍□○]+([\d.\-년월일\s~()가-힣]+)",
        content_text
    )
    if m:
        period = clean(m.group(2))[:80]

    # ── support
    support: list[str] = []
    if re.search(r"무료|참가비\s*없음|수강료\s*없음", content_text):
        support.append("무료")
    m = re.search(
        r"(지원\s*내용|지원\s*사항|주요\s*내용)[^\n:❍□○]*[:\s❍□○]+([\s\S]{10,400}?)(?=\n\n|❍|□|\Z)",
        content_text
    )
    if m:
        items = re.split(r"[\n,◦]", m.group(2))
        support += [clean(i) for i in items if 4 < len(clean(i)) < 100]
    support = list(dict.fromkeys(support))[:6]

    # ── duration
    duration = period
    m = re.search(r"(총\s*\d+회|\d+개월|\d+주|매주)", content_text)
    if m:
        duration = (duration + " / " + m.group()).strip(" /")

    # ── spots
    spots: int | None = None
    m = re.search(r"(\d+)\s*명\s*(이내|모집|선발|내외|정원)", content_text)
    if m:
        spots = int(m.group(1))

    # ── cover: 본문 첫 이미지
    cover = ""
    if content:
        img = content.select_one("img[src*='fileManager'], img[src*='bbs']")
        if img:
            src = img.get("src", "")
            cover = urljoin(BASE_URL, src) if src.startswith("/") else src

    # ── tag
    tag = category or ""
    if not tag:
        for kw in RELEVANT_KEYWORDS:
            if kw.lower() in title.lower():
                tag = kw
                break

    # ── summary
    summary = ""
    for sent in re.split(r"[\n。]", content_text):
        s = clean(sent)
        if len(s) > 20:
            summary = s[:120]
            break

    # ── detail
    detail = content_text[:600]

    # ── bullets: ❍ / □ / ○ / ▶ 항목
    bullets: list[str] = []
    for mb in re.finditer(r"[❍□○▶◇•]\s*([^\n❍□○▶◇•]{5,80})", content_text):
        b = clean(mb.group(1))
        if b and b not in bullets:
            bullets.append(b)
    if not bullets:
        for mb in re.finditer(r"\d+[..]\s*([^\n]{5,80})", content_text):
            b = clean(mb.group(1))
            if b and b not in bullets:
                bullets.append(b)
    bullets = bullets[:6]

    return {
        "title":     title,
        "org":       org,
        "region":    region,
        "mode":      mode,
        "motive":    motive,
        "status":    status,
        "period":    period,
        "support":   support,
        "deadline":  deadline,
        "duration":  duration,
        "spots":     spots,
        "spotsLeft": None,
        "cover":     cover,
        "tag":       tag,
        "summary":   summary,
        "detail":    detail,
        "bullets":   bullets,
        "pub_date":  pub_date,
        "url":       url,
    }


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

COLUMNS = [
    "title", "org", "region", "mode", "motive", "status",
    "period", "support", "deadline", "duration", "spots",
    "spotsLeft", "cover", "tag", "summary", "detail", "bullets",
    "pub_date", "url",
]
COL_WIDTHS = {
    "title": 38, "org": 20, "region": 10, "mode": 12, "motive": 35,
    "status": 14, "period": 28, "support": 35, "deadline": 20,
    "duration": 28, "spots": 10, "spotsLeft": 10, "cover": 30,
    "tag": 14, "summary": 42, "detail": 52, "bullets": 42,
    "pub_date": 14, "url": 32,
}
CENTER_COLS = {"region", "mode", "status", "spots", "spotsLeft", "tag", "pub_date"}


def save_excel(results: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공지사항"

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    side   = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    odd    = PatternFill("solid", start_color="EBF3FA")
    even   = PatternFill("solid", start_color="FFFFFF")

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 25

    for ri, ev in enumerate(results, 2):
        fill = odd if ri % 2 == 1 else even
        for ci, col in enumerate(COLUMNS, 1):
            val = ev.get(col)
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in CENTER_COLS else left
            if col == "url" and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
    print(f"  ✅ 엑셀 저장 완료 → {path}")


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  충북청년희망센터 공지사항 관련 게시물 스크래퍼")
    print(f"  필터 키워드: {', '.join(RELEVANT_KEYWORDS[:6])} 외 {len(RELEVANT_KEYWORDS)-6}개")
    print("=" * 60)

    print("\n[1/3] 전체 목록에서 게시물 링크 수집 중...")
    hints = get_post_links()

    print("\n[2/3] 상세 정보 추출 + 관련성 필터링 중...")
    results = []
    skipped = 0
    for i, hint in enumerate(hints, 1):
        print(f"  ({i:02d}/{len(hints)}) {hint['title'][:50]}...")
        detail = parse_detail(hint)
        if detail:
            results.append(detail)
        else:
            skipped += 1
            print(f"    → 관련성 없음, skip")
        time.sleep(random.uniform(0.4, 0.9))

    print(f"\n  → 최종 {len(results)}개 수집 / {skipped}개 skip")

    if not results:
        print("  ⚠️  관련 게시물을 찾지 못했습니다.")
        return

    print("\n[3/3] 저장 중...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"cbhope_events_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 저장 완료 → {json_path}")

    save_excel(results, f"cbhope_events_{ts}.xlsx")

    print("\n" + "=" * 60)
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print("=" * 60)
    print(f"\n✅ 완료! 총 {len(results)}개 관련 게시물")


if __name__ == "__main__":
    main()
