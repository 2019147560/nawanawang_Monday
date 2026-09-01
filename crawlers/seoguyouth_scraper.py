"""
서청플 서구 청년정보플랫폼(seoguyouth.kr) 프로그램 스크래퍼
- /program/program_list/3 전체 수집 (142건 / 12페이지)
- 상세 페이지에서 정형화된 필드 정확히 파싱
사용법: python seoguyouth_scraper.py
결과물: seoguyouth_events_YYYYMMDD_HHMMSS.json / .xlsx
"""

import re
import time
import json
import random
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

BASE_URL  = "https://www.seoguyouth.kr"
LIST_URL  = "https://www.seoguyouth.kr/program/program_list/3"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": BASE_URL,
}

ORG_NAME = "서구 청년정보플랫폼 (청춘스럽)"


# ── 유틸 ─────────────────────────────────────────────────────────────────────

def fetch(url: str) -> BeautifulSoup | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [오류] {url} → {e}")
        return None


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


# ── 1단계: 목록 페이지에서 링크 수집 ─────────────────────────────────────────

def get_post_links() -> list[dict]:
    """
    목록 구조:
    - li > a[href*='/program/program_view/']
    - 링크 텍스트에 제목, 상태(모집중/모집종료), 운영기간, 해시태그 포함
    """
    links = []
    seen  = set()

    for page in range(1, 20):
        url = f"{LIST_URL}?page={page}"
        print(f"  목록 페이지 {page} 파싱 중...")
        soup = fetch(url)
        if not soup:
            break

        anchors = soup.select("a[href*='/program/program_view/']")
        found_any = False

        for a in anchors:
            href = a.get("href", "")
            # page 파라미터 없이 고유 URL로 정규화
            m = re.search(r"/program/program_view/(\d+)", href)
            if not m:
                continue
            prog_id = m.group(1)
            if prog_id in seen:
                continue

            full_url = f"{BASE_URL}/program/program_view/{prog_id}?page={page}"
            text = clean(a.get_text(" "))

            # 목록에서 미리 파악 가능한 정보 추출
            # 상태: 모집중 / 모집종료
            status_hint = ""
            if "모집중" in text:
                status_hint = "모집중"
            elif "모집종료" in text:
                status_hint = "모집종료"

            # 제목: 상태 다음 텍스트 패턴 파악
            title_m = re.search(r"모집(?:중|종료)\s+(.+?)\s+(?:운영기간|#|\Z)", text)
            title_hint = clean(title_m.group(1)) if title_m else ""

            # 해시태그
            tags = re.findall(r"#([^\s#]+)", text)

            # 썸네일 이미지
            img = a.select_one("img[src*='uploads']")
            cover = urljoin(BASE_URL, img.get("src", "")) if img else ""

            seen.add(prog_id)
            links.append({
                "url":         full_url,
                "prog_id":     prog_id,
                "title":       title_hint,
                "status_hint": status_hint,
                "tags":        tags,
                "cover":       cover,
                "list_page":   page,
            })
            found_any = True

        if not found_any:
            print(f"  → 페이지 {page}에 게시물 없음, 중단")
            break

        # 마지막 페이지 확인
        next_link = soup.select_one(f"a[href*='page={page+1}']")
        if not next_link:
            print(f"  → 마지막 페이지 도달 ({page})")
            break

        time.sleep(random.uniform(0.4, 0.8))

    return links


# ── 2단계: 상세 페이지 파싱 ──────────────────────────────────────────────────

def parse_detail(hint: dict) -> dict:
    url  = hint["url"]
    soup = fetch(url)
    if not soup:
        return {}

    # ── title: h2 또는 상세 제목 영역
    # 구조: 상태배지 + 썸네일 img[alt] + 해시태그 + 제목 텍스트
    title = ""
    title_tag = soup.select_one(".program_title, .view_title, h2, h3")
    if title_tag:
        title = clean(title_tag.get_text())
    # fallback: img alt (썸네일 alt에 제목이 있음)
    if not title:
        thumb = soup.select_one("img[src*='uploads/program']")
        if thumb:
            title = clean(thumb.get("alt", ""))
    if not title:
        title = hint.get("title", "")

    # ── 정형화된 필드 파싱: dl > dt + dd 또는 ul > li > strong + span
    # 실제 구조: li > strong(라벨) + 텍스트
    field_map: dict[str, str] = {}
    for li in soup.select("ul li, dl dt, .info_list li"):
        strong = li.select_one("strong, dt")
        if strong:
            label = clean(strong.get_text())
            # strong 다음 텍스트 노드
            val = clean(li.get_text().replace(label, "", 1))
            field_map[label] = val

    # 직접 텍스트 패턴으로도 보완
    full_text = soup.get_text(" ")

    def get_field(*keys) -> str:
        for k in keys:
            if k in field_map and field_map[k]:
                return field_map[k]
        # 텍스트 패턴 fallback
        for k in keys:
            m = re.search(rf"{k}\s*([^\n운모문선정]{3,60})", full_text)
            if m:
                return clean(m.group(1))
        return ""

    # ── 운영기간 (period)
    period = get_field("운영기간")
    period = re.sub(r"\s*00:00", "", period).strip()

    # ── 모집기간 (deadline)
    deadline_raw = get_field("모집기간")
    deadline_raw = re.sub(r"\s*00:00", "", deadline_raw).strip()
    # 마감일만 추출 (~ 이후)
    deadline = ""
    if "~" in deadline_raw:
        deadline = clean(deadline_raw.split("~")[-1])
    else:
        deadline = deadline_raw

    # ── 모집대상 (motive)
    motive_raw = get_field("모집대상")
    motive = [clean(i) for i in re.split(r"[,、\n]", motive_raw) if len(clean(i)) > 3][:5]

    # ── 모집인원 / 현재인원 (spots / spotsLeft)
    spots: int | None = None
    spots_left: int | None = None
    capacity_raw = get_field("모집인원")
    # 패턴: "2 / 6명" 또는 "6명"
    cap_m = re.search(r"(\d+)\s*/\s*(\d+)\s*명", capacity_raw)
    if cap_m:
        spots_left = int(cap_m.group(1))   # 현재 신청 인원
        spots      = int(cap_m.group(2))   # 총 정원
        spots_left = spots - spots_left    # 남은 인원
    else:
        num_m = re.search(r"(\d+)\s*명", capacity_raw)
        if num_m:
            spots = int(num_m.group(1))

    # ── 문의전화 (contact)
    contact = get_field("문의전화", "문의")

    # ── 운영장소
    place = get_field("운영장소", "장소")

    # ── 운영요일
    weekday = get_field("운영요일", "요일")

    # ── 선정방식
    selection = get_field("선정방식")

    # ── 상태
    status_hint = hint.get("status_hint", "")
    if status_hint == "모집중":
        status = "현재 신청 가능"
    elif status_hint == "모집종료":
        status = "마감"
    else:
        # 페이지 내 상태 배지
        badge = soup.select_one(".status, .badge, .state")
        badge_text = clean(badge.get_text()) if badge else ""
        if "모집중" in badge_text or "모집중" in full_text[:200]:
            status = "현재 신청 가능"
        elif "모집종료" in badge_text or "모집종료" in full_text[:200]:
            status = "마감"
        else:
            # 마감일 기준 판단
            now = datetime.now()
            dm = re.search(r"(20\d{2})[.\s]\s*(\d{2})[.\s]\s*(\d{2})", deadline)
            if dm:
                try:
                    dl = datetime(int(dm.group(1)), int(dm.group(2)), int(dm.group(3)))
                    status = "마감" if now > dl else "현재 신청 가능"
                except Exception:
                    status = "현재 신청 가능"
            else:
                status = "현재 신청 가능"

    # ── org: 청춘스럽 고정 (breadcrumb에서 확인)
    org = ORG_NAME
    breadcrumb = soup.select_one(".breadcrumb, .location")
    if breadcrumb:
        org_text = clean(breadcrumb.get_text())
        if "청춘정거장" in org_text:
            org = "서구 청년정보플랫폼 (청춘정거장)"
        elif "청춘포털" in org_text:
            org = "서구 청년정보플랫폼 (청춘포털)"
        elif "공간 캘린더" in org_text:
            org = "서구 청년정보플랫폼 (공간 캘린더)"

    # ── region
    region = "대전"
    if place and any(r in place for r in ["서구", "동구", "중구", "유성", "대덕"]):
        region = f"대전 {re.search(r'[가-힣]+구', place).group()}"

    # ── mode
    if place and re.search(r"온라인|zoom|유튜브", place, re.I):
        mode = "온라인"
    elif place:
        mode = "오프라인"
    else:
        mode = ""

    # ── tag: 해시태그
    tag_list = hint.get("tags", [])
    if not tag_list:
        tag_list = re.findall(r"#([^\s#]+)", full_text)
    tag = " ".join(f"#{t}" for t in tag_list[:5])

    # ── cover: 썸네일 우선
    cover = hint.get("cover", "")
    if not cover:
        thumb = soup.select_one("img[src*='uploads/program']")
        if thumb:
            src = thumb.get("src", "")
            cover = urljoin(BASE_URL, src) if src.startswith("/") else src

    # ── 본문 (에디터 영역)
    content = soup.select_one(".editor_view, .view_content, .program_content, article")
    content_text = clean(content.get_text(" ")) if content else ""

    # ── summary: 본문 첫 의미있는 문장
    summary = ""
    for sent in re.split(r"[\n。]", content_text):
        s = clean(sent)
        if len(s) > 15:
            summary = s[:120]
            break

    # ── detail
    detail = content_text[:600] if content_text else full_text[:600]

    # ── support: 무료 여부 + 지원내용
    support: list[str] = []
    if re.search(r"무료|참가비\s*없음|수강료\s*없음", full_text):
        support.append("무료")
    if place:
        support.append(f"장소: {place}")
    if weekday:
        support.append(f"운영요일: {weekday}")
    if selection:
        support.append(f"선정방식: {selection}")
    support = support[:5]

    # ── duration
    duration = period
    m = re.search(r"(총\s*\d+회|\d+주|\d+개월|매주|매월)", full_text)
    if m:
        duration = (duration + " / " + m.group()).strip(" /")

    # ── bullets: 본문 내 항목
    bullets: list[str] = []
    if content:
        for mb in re.finditer(r"[•·▶○◆■]\s*([^\n•·▶○◆■]{5,80})", content_text):
            b = clean(mb.group(1))
            if b and b not in bullets:
                bullets.append(b)
    if not bullets and tag_list:
        bullets = [f"#{t}" for t in tag_list[:6]]
    bullets = bullets[:6]

    return {
        "title":     title,
        "org":       org,
        "region":    region,
        "mode":      mode,
        "motive":    motive,
        "status":    status,
        "period":    period,
        "support":   support,
        "deadline":  deadline,
        "duration":  duration,
        "spots":     spots,
        "spotsLeft": spots_left,
        "cover":     cover,
        "tag":       tag,
        "summary":   summary,
        "detail":    detail,
        "bullets":   bullets,
        "contact":   contact,
        "url":       url,
    }


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

COLUMNS = [
    "title", "org", "region", "mode", "motive", "status",
    "period", "support", "deadline", "duration", "spots",
    "spotsLeft", "cover", "tag", "summary", "detail", "bullets",
    "contact", "url",
]
COL_WIDTHS = {
    "title": 38, "org": 22, "region": 10, "mode": 10, "motive": 30,
    "status": 14, "period": 28, "support": 28, "deadline": 20,
    "duration": 28, "spots": 8, "spotsLeft": 10, "cover": 28,
    "tag": 30, "summary": 42, "detail": 50, "bullets": 38,
    "contact": 18, "url": 32,
}
CENTER_COLS = {"region", "mode", "status", "spots", "spotsLeft"}


def save_excel(results: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "프로그램 목록"

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    side   = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    odd    = PatternFill("solid", start_color="EBF3FA")
    even   = PatternFill("solid", start_color="FFFFFF")

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 25

    for ri, ev in enumerate(results, 2):
        fill = odd if ri % 2 == 1 else even
        for ci, col in enumerate(COLUMNS, 1):
            val = ev.get(col)
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in CENTER_COLS else left
            if col == "url" and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
    print(f"  ✅ 엑셀 저장 완료 → {path}")


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  서청플 청춘스럽 프로그램 스크래퍼")
    print("=" * 60)

    print("\n[1/3] 목록에서 프로그램 링크 수집 중...")
    hints = get_post_links()
    print(f"  → {len(hints)}개 프로그램 발견")

    if not hints:
        print("  ⚠️  프로그램을 찾지 못했습니다.")
        return

    print("\n[2/3] 상세 정보 추출 중...")
    results = []
    for i, hint in enumerate(hints, 1):
        print(f"  ({i:03d}/{len(hints)}) [{hint['status_hint']}] {hint['title'][:45]}...")
        detail = parse_detail(hint)
        if detail:
            results.append(detail)
        time.sleep(random.uniform(0.4, 0.8))

    print("\n[3/3] 저장 중...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"seoguyouth_events_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 저장 완료 → {json_path}")

    save_excel(results, f"seoguyouth_events_{ts}.xlsx")

    print("\n" + "=" * 60)
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print("=" * 60)
    print(f"\n✅ 완료! 총 {len(results)}개 프로그램")


if __name__ == "__main__":
    main()
