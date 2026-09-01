"""
사단법인 늘봄청소년(springyouth.mixon.io) 공지사항 스크래퍼
- mixon.io Rails/Turbo 기반: 목록은 Selenium, 상세는 requests
- 전체 게시물 수집 (총 13건)
사용법: python springyouth_scraper.py
결과물: springyouth_events_YYYYMMDD_HHMMSS.json / .xlsx

필요 패키지:
  pip install requests beautifulsoup4 openpyxl selenium webdriver-manager
"""

import re
import time
import json
import random
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

BASE_URL  = "https://springyouth.mixon.io"
LIST_URL  = "https://springyouth.mixon.io/boards/notice"
PAGE_URL  = "https://springyouth.mixon.io/boards/notice?page={page}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": BASE_URL,
}

ORG_NAME = "사단법인 늘봄청소년"


# ── 유틸 ─────────────────────────────────────────────────────────────────────

def fetch(url: str) -> BeautifulSoup | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [오류] {url} → {e}")
        return None


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


# ── 1단계: Selenium으로 목록 수집 ────────────────────────────────────────────

def get_post_links_selenium() -> list[dict]:
    """
    mixon.io 목록은 JS 렌더링 → Selenium으로 /posts/{slug} 링크 수집
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
    except ImportError:
        print("  [오류] selenium 또는 webdriver-manager가 설치되지 않았습니다.")
        print("  pip install selenium webdriver-manager")
        return []

    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument(f"user-agent={HEADERS['User-Agent']}")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts
    )

    links = []
    seen  = set()

    try:
        for page in range(1, 10):
            url = PAGE_URL.format(page=page)
            print(f"  [Selenium] 목록 페이지 {page} 로드 중...")
            driver.get(url)

            # 게시물 링크 로딩 대기
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/posts/']"))
                )
            except Exception:
                # 링크가 없으면 마지막 페이지
                print(f"  → 페이지 {page}에 게시물 없음, 중단")
                break

            # /posts/{slug} 패턴 링크 수집
            anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='/posts/']")
            found_any = False

            for a in anchors:
                href = a.get_attribute("href") or ""
                if not re.search(r"/posts/[A-Za-z0-9]+$", href):
                    continue
                if href in seen:
                    continue

                title = clean(a.text)
                # 제목 없으면 부모 요소에서 추출
                if not title or len(title) < 2:
                    try:
                        parent = a.find_element(By.XPATH, "..")
                        title = clean(parent.text)
                    except Exception:
                        pass

                # 썸네일
                cover = ""
                try:
                    img = a.find_element(By.TAG_NAME, "img")
                    cover = img.get_attribute("src") or ""
                except Exception:
                    pass

                seen.add(href)
                links.append({"url": href, "title": title, "cover": cover})
                found_any = True

            if not found_any:
                print(f"  → 페이지 {page}에 게시물 없음, 중단")
                break

            time.sleep(random.uniform(0.5, 1.0))

    finally:
        driver.quit()

    return links


# ── 1단계 대안: requests만으로 목록 수집 시도 ────────────────────────────────
# mixon.io는 Accept: application/json 헤더로 JSON 응답을 주기도 함

def get_post_links_requests() -> list[dict]:
    """
    Selenium 없이 시도: Accept: application/json 요청
    실패 시 빈 리스트 반환
    """
    links = []
    seen  = set()

    for page in range(1, 10):
        url = PAGE_URL.format(page=page)
        try:
            r = requests.get(
                url,
                headers={**HEADERS, "Accept": "application/json, text/javascript, */*"},
                timeout=15
            )
            # JSON 응답이면 파싱
            if "application/json" in r.headers.get("Content-Type", ""):
                data = r.json()
                posts = data.get("posts", data.get("records", data.get("items", [])))
                if not posts:
                    break
                for post in posts:
                    slug = post.get("slug") or post.get("id", "")
                    title = post.get("title", "")
                    full_url = f"{BASE_URL}/posts/{slug}"
                    if full_url not in seen:
                        seen.add(full_url)
                        links.append({"url": full_url, "title": title, "cover": ""})
            else:
                # HTML 응답 → BeautifulSoup으로 파싱 시도
                soup = BeautifulSoup(r.text, "html.parser")
                anchors = soup.select("a[href*='/posts/']")
                found = False
                for a in anchors:
                    href = a.get("href", "")
                    if not re.search(r"/posts/[A-Za-z0-9]+$", href):
                        continue
                    full_url = urljoin(BASE_URL, href)
                    if full_url in seen:
                        continue
                    seen.add(full_url)
                    links.append({
                        "url":   full_url,
                        "title": clean(a.get_text()),
                        "cover": "",
                    })
                    found = True
                if not found:
                    break
        except Exception as e:
            print(f"  [경고] 페이지 {page} 요청 실패: {e}")
            break

        time.sleep(random.uniform(0.4, 0.8))

    return links


def get_post_links() -> list[dict]:
    """requests 먼저 시도, 실패 시 Selenium"""
    print("  [방법1] requests로 목록 수집 시도...")
    links = get_post_links_requests()
    if links:
        print(f"  → requests 성공! {len(links)}개 링크")
        return links

    print("  [방법2] Selenium으로 목록 수집 시도...")
    links = get_post_links_selenium()
    return links


# ── 2단계: 상세 페이지 파싱 ──────────────────────────────────────────────────

def parse_detail(hint: dict) -> dict:
    url  = hint["url"]
    soup = fetch(url)
    if not soup:
        return {}

    # ── title: h2 태그
    h2 = soup.select_one("h2")
    title = clean(h2.get_text()) if h2 else hint.get("title", "")
    # " | 사단법인 늘봄청소년" 접미사 제거
    title = re.sub(r"\s*\|\s*사단법인.*$", "", title).strip()

    # ── 발행일: "발행일 YYYY-MM-DD" 텍스트
    pub_date = ""
    dm = re.search(r"발행일\s*(20\d{2}-\d{2}-\d{2})", soup.get_text())
    if dm:
        pub_date = dm.group(1)

    # ── 본문
    content = (
        soup.select_one(".post-content") or
        soup.select_one(".fr-view") or
        soup.select_one("article") or
        soup.select_one(".prose")
    )
    if not content:
        h2_tag = soup.find("h2")
        content = h2_tag.find_next("div") if h2_tag else None

    content_text = clean(content.get_text(" ")) if content else ""

    # ── org
    org = ORG_NAME
    m = re.search(r"(주최|주관|운영)[:\s]*([^\n,/]{2,30})", content_text)
    if m:
        org = clean(m.group(2))

    # ── region
    region = "강원"  # 법인 본부 기준 (강원도)
    region_list = ["전국", "서울", "부산", "대구", "인천", "광주", "대전", "울산",
                   "세종", "경기", "충북", "충남", "전북", "전남", "경북", "경남",
                   "제주", "춘천"]
    for r in region_list:
        if r in content_text:
            region = r
            break

    # ── mode
    if re.search(r"온·오프|온\s*오프|혼합", content_text):
        mode = "혼합"
    elif re.search(r"온라인|비대면|zoom|유튜브|화상|카카오", content_text, re.I):
        mode = "온라인"
    elif re.search(r"오프라인|현장|대면|방문|센터|기숙", content_text):
        mode = "오프라인"
    else:
        mode = ""

    # ── motive: 대상 항목
    motive: list[str] = []
    m = re.search(r"(대\s*상|신청자격|참여자격)[^\n:]*[:\s]+([\s\S]{5,200}?)(?=\n\n|○|▶|◇|\Z)", content_text)
    if m:
        block = m.group(2)
        items = re.split(r"[,、\n]", block)
        motive = [clean(i) for i in items if 3 < len(clean(i)) < 80][:5]

    # ── deadline
    deadline = ""
    m = re.search(r"(신청\s*마감|모집\s*기간|접수)[^\d]*([\d.\-년월일\s~]+)", content_text)
    if m:
        deadline = clean(m.group(2))

    # ── status
    if re.search(r"마감|종료|접수종료|모집완료", content_text[:300]):
        status = "마감"
    elif re.search(r"예정|추후공지|미정", content_text[:300]):
        status = "모집 예정"
    else:
        now = datetime.now()
        date_m = re.search(r"(20\d{2})[.\-년]\s*(\d{1,2})[.\-월]\s*(\d{1,2})", deadline)
        if date_m:
            try:
                dl = datetime(int(date_m.group(1)), int(date_m.group(2)), int(date_m.group(3)))
                status = "마감" if now > dl else "현재 신청 가능"
            except Exception:
                status = "현재 신청 가능"
        else:
            status = "현재 신청 가능"

    # ── period
    period = ""
    m = re.search(r"(일\s*시|기\s*간|운영\s*기간|프로그램\s*기간)[^\n:]*[:\s]+([\d.\-년월일(가-힣)\s~]+)", content_text)
    if m:
        period = clean(m.group(2))[:80]

    # ── support
    support: list[str] = []
    if re.search(r"무료|참가비\s*없음|수강료\s*없음", content_text):
        support.append("무료")
    m = re.search(r"(지원|혜택|제공)[^\n:]*[:\s]+([\s\S]{5,200}?)(?=\n\n|○|▶|\Z)", content_text)
    if m:
        items = re.split(r"[\n,]", m.group(2))
        support += [clean(i) for i in items if 4 < len(clean(i)) < 80]
    support = list(dict.fromkeys(support))[:5]

    # ── duration
    duration = period
    m = re.search(r"(총\s*\d+회|\d+박\s*\d+일|\d+주|\d+개월|매주)", content_text)
    if m:
        duration = (duration + " / " + m.group()).strip(" /")

    # ── spots
    spots: int | None = None
    m = re.search(r"(\d+)\s*명\s*(이내|모집|선발|내외|정원)", content_text)
    if m:
        spots = int(m.group(1))

    # ── cover: 힌트 or 본문 첫 이미지
    cover = hint.get("cover", "")
    if not cover and content:
        img = content.select_one("img[src*='storage'], img[src*='active_storage']")
        if img:
            src = img.get("src", "")
            cover = urljoin(BASE_URL, src) if src.startswith("/") else src

    # ── tag: 제목 앞 [] 태그
    tag = ""
    tm = re.search(r"^\[([^\]]{1,15})\]", title)
    if tm:
        tag = tm.group(1)

    # ── summary
    summary = ""
    if content:
        bold = content.select_one("strong, b")
        if bold:
            s = clean(bold.get_text())
            if len(s) > 10:
                summary = s[:120]
    if not summary:
        for sent in re.split(r"[\n。]", content_text):
            s = clean(sent)
            if len(s) > 15:
                summary = s[:120]
                break

    # ── detail
    detail = content_text[:600]

    # ── bullets: ○ / ▶ / • / - 항목
    bullets: list[str] = []
    for mb in re.finditer(r"[○▶◇•]\s*([^\n○▶◇•]{5,80})", content_text):
        b = clean(mb.group(1))
        if b and b not in bullets:
            bullets.append(b)
    if not bullets:
        for mb in re.finditer(r"-\s+([^\n\-]{5,80})", content_text):
            b = clean(mb.group(1))
            if b and b not in bullets:
                bullets.append(b)
    bullets = bullets[:6]

    return {
        "title":     title,
        "org":       org,
        "region":    region,
        "mode":      mode,
        "motive":    motive,
        "status":    status,
        "period":    period,
        "support":   support,
        "deadline":  deadline,
        "duration":  duration,
        "spots":     spots,
        "spotsLeft": None,
        "cover":     cover,
        "tag":       tag,
        "summary":   summary,
        "detail":    detail,
        "bullets":   bullets,
        "pub_date":  pub_date,
        "url":       url,
    }


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

COLUMNS = [
    "title", "org", "region", "mode", "motive", "status",
    "period", "support", "deadline", "duration", "spots",
    "spotsLeft", "cover", "tag", "summary", "detail", "bullets",
    "pub_date", "url",
]
COL_WIDTHS = {
    "title": 38, "org": 20, "region": 10, "mode": 12, "motive": 35,
    "status": 14, "period": 28, "support": 30, "deadline": 18,
    "duration": 28, "spots": 10, "spotsLeft": 10, "cover": 30,
    "tag": 14, "summary": 42, "detail": 52, "bullets": 42,
    "pub_date": 14, "url": 32,
}
CENTER_COLS = {"region", "mode", "status", "spots", "spotsLeft", "tag", "pub_date"}


def save_excel(results: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공지사항"

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    side   = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    odd    = PatternFill("solid", start_color="EBF3FA")
    even   = PatternFill("solid", start_color="FFFFFF")

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 25

    for ri, ev in enumerate(results, 2):
        fill = odd if ri % 2 == 1 else even
        for ci, col in enumerate(COLUMNS, 1):
            val = ev.get(col)
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in CENTER_COLS else left
            if col == "url" and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
    print(f"  ✅ 엑셀 저장 완료 → {path}")


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  사단법인 늘봄청소년(springyouth.mixon.io) 스크래퍼")
    print("=" * 60)

    print("\n[1/3] 목록에서 게시물 링크 수집 중...")
    hints = get_post_links()
    print(f"  → {len(hints)}개 게시물 발견")

    if not hints:
        print("  ⚠️  게시물을 찾지 못했습니다.")
        print("  💡 Selenium 설치 여부 확인: pip install selenium webdriver-manager")
        return

    print("\n[2/3] 상세 정보 추출 중...")
    results = []
    for i, hint in enumerate(hints, 1):
        print(f"  ({i:02d}/{len(hints)}) {hint['title'][:50]}...")
        detail = parse_detail(hint)
        if detail:
            results.append(detail)
        time.sleep(random.uniform(0.5, 1.0))

    print("\n[3/3] 저장 중...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"springyouth_events_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 저장 완료 → {json_path}")

    save_excel(results, f"springyouth_events_{ts}.xlsx")

    print("\n" + "=" * 60)
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print("=" * 60)
    print(f"\n✅ 완료! 총 {len(results)}개 게시물")


if __name__ == "__main__":
    main()
