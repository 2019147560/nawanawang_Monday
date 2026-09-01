"""
대구광역시 청년센터(dgyouth.kr) 공지사항 스크래퍼
- notice.asp 전체 페이지 순회 (총 133페이지)
- ▶ 기호 구조화된 본문 정확히 파싱
사용법: python dgyouth_scraper.py
결과물: dgyouth_events_YYYYMMDD_HHMMSS.json / .xlsx
"""

import re
import time
import json
import random
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

BASE_URL = "http://www.dgyouth.kr"
LIST_URL = "http://www.dgyouth.kr/board/notice.asp"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": BASE_URL,
}

ORG_NAME = "대구광역시 청년센터"

# 공간별 연락처 매핑
SPACE_CONTACT = {
    "공감그래": "053-427-1934",
    "활동그래": "053-426-1939",
    "행복그래": "053-426-1934",
    "다온나그래": "053-427-1938",
}


# ── 유틸 ─────────────────────────────────────────────────────────────────────

def fetch(url: str) -> BeautifulSoup | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [오류] {url} → {e}")
        return None


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


# ── 1단계: 목록 페이지에서 링크 수집 ─────────────────────────────────────────

def get_post_links(max_pages: int = 133) -> list[dict]:
    """
    목록 구조:
    - li > a[href*='num='] (카드형 목록)
    - 링크 텍스트에 접수기간/진행기간 포함
    - 썸네일: li > a > img[src*='/pds/notice/']
    """
    links = []
    seen  = set()

    for page in range(1, max_pages + 1):
        url = f"{LIST_URL}?page={page}&sKey=&sWord=&pCate=0"
        print(f"  목록 페이지 {page}/{max_pages} 파싱 중...")
        soup = fetch(url)
        if not soup:
            break

        anchors = soup.select("a[href*='num='][href*='pmode=VIEW']")
        found_any = False

        for a in anchors:
            href = a.get("href", "")
            m = re.search(r"num=(\d+)", href)
            if not m:
                continue
            num = m.group(1)
            if num in seen:
                continue

            full_url = urljoin(BASE_URL, href)
            text = clean(a.get_text(" "))

            # 제목 추출: [공지] [센터사업] 제목 접수기간... 패턴
            title_m = re.search(r"(?:\[공지\]\s*)?(?:\[[^\]]+\]\s*)?(.+?)(?:\s+접수기간|\s+진행기간|\Z)", text)
            title_hint = clean(title_m.group(1)) if title_m else text[:60]

            # 접수기간
            deadline_m = re.search(r"접수기간\s*:\s*([^\n진행]+)", text)
            deadline_hint = clean(deadline_m.group(1)) if deadline_m else ""

            # 진행기간
            period_m = re.search(r"진행기간\s*:\s*([^\n]+)", text)
            period_hint = clean(period_m.group(1)) if period_m else ""

            # 썸네일
            img = a.select_one("img[src*='/pds/notice/']")
            cover = urljoin(BASE_URL, img.get("src", "")) if img else ""

            # 카테고리 [센터사업] / [공고] 등
            cat_m = re.search(r"\[(?!공지)([^\]]+)\]", text)
            category = cat_m.group(1) if cat_m else ""

            seen.add(num)
            links.append({
                "url":           full_url,
                "num":           num,
                "title":         title_hint,
                "deadline_hint": deadline_hint,
                "period_hint":   period_hint,
                "category":      category,
                "cover":         cover,
                "list_page":     page,
            })
            found_any = True

        if not found_any:
            print(f"  → 페이지 {page}에 게시물 없음, 중단")
            break

        # 마지막 페이지 확인
        next_link = soup.find("a", href=re.compile(rf"page={page+1}"))
        if not next_link and page > 1:
            print(f"  → 마지막 페이지 도달 ({page})")
            break

        time.sleep(random.uniform(0.4, 0.8))

    return links


# ── 2단계: 상세 페이지 파싱 ──────────────────────────────────────────────────

def parse_detail(hint: dict) -> dict:
    url  = hint["url"]
    soup = fetch(url)
    if not soup:
        return {}

    # ── title: 본문 상단 굵은 제목
    title = ""
    # [공지] 태그 다음 텍스트 또는 h 태그
    for tag in soup.select("strong, b, h2, h3, .view-title"):
        t = clean(tag.get_text())
        if len(t) > 5 and "[공지]" not in t and "메뉴" not in t:
            title = t
            break
    if not title:
        title = hint.get("title", "")

    # ── 등록일
    pub_date = ""
    dm = re.search(r"등록일자\s*(20\d{2}\.\d{2}\.\d{2})", soup.get_text())
    if dm:
        pub_date = dm.group(1).replace(".", "-")

    # ── 본문 영역 (content div)
    content = soup.select_one("#content, .board-view-content, .view_content, .content")
    if not content:
        # 목록으로 돌아가기 링크 전까지
        body = soup.find("body")
        content = body

    content_text = clean(content.get_text(" ")) if content else ""

    # ── ▶ 기호 기반 필드 파싱
    def get_arrow_field(*keys) -> str:
        for k in keys:
            m = re.search(
                rf"▶\s*{k}\s*[:\s]*([^\n▶☎]{3,200})",
                content_text
            )
            if m:
                return clean(m.group(1))
        return ""

    # ── 일시/진행기간 (period)
    period = get_arrow_field("일시", "진행기간", "기간") or hint.get("period_hint", "")

    # ── 접수기간 (deadline)
    deadline_raw = get_arrow_field("접수기간", "신청기간", "모집기간") or hint.get("deadline_hint", "")
    # 마감일만 추출
    deadline = ""
    if "~" in deadline_raw:
        deadline = clean(deadline_raw.split("~")[-1])
    else:
        deadline = deadline_raw

    # ── 대상 (motive)
    motive: list[str] = []
    target_raw = get_arrow_field("대상", "참여대상", "신청대상")
    if target_raw:
        items = re.split(r"[-\n·]", target_raw)
        motive = [clean(i) for i in items if 3 < len(clean(i)) < 100][:5]

    # ── 내용 (support / bullets)
    content_raw = get_arrow_field("내용", "주요일정", "프로그램")

    # ── 장소
    place = get_arrow_field("장소")

    # ── 참여비용
    cost = get_arrow_field("참여비용", "비용", "수강료")

    # ── 문의
    contact = ""
    tel_m = re.search(r"☎\s*문의[:\s]*([^\n]+)", content_text)
    if tel_m:
        contact = clean(tel_m.group(1))
    else:
        tel_m = re.search(r"문의[:\s]*([^\n]+)", content_text)
        if tel_m:
            contact = clean(tel_m.group(1))

    # ── org: 공간명에 따라 세분화
    org = ORG_NAME
    for space, tel in SPACE_CONTACT.items():
        if space in content_text or space in hint.get("category", ""):
            org = f"{ORG_NAME} ({space})"
            if not contact:
                contact = tel
            break

    # ── region
    region = "대구"
    if place:
        regions = ["중구", "동구", "서구", "남구", "북구", "수성구", "달서구", "달성군"]
        for r in regions:
            if r in place:
                region = f"대구 {r}"
                break

    # ── mode
    if re.search(r"온라인|zoom|유튜브|비대면", content_text, re.I):
        mode = "온라인"
    elif re.search(r"온·오프|혼합", content_text):
        mode = "혼합"
    elif place:
        mode = "오프라인"
    else:
        mode = ""

    # ── status
    if re.search(r"마감|종료|모집완료", content_text[:300] + hint.get("title", "")):
        status = "마감"
    elif re.search(r"예정|추후공지", content_text[:300]):
        status = "모집 예정"
    else:
        now = datetime.now()
        dl_m = re.search(r"(20\d{2})[.\s](\d{1,2})[.\s](\d{1,2})", deadline)
        if dl_m:
            try:
                dl = datetime(int(dl_m.group(1)), int(dl_m.group(2)), int(dl_m.group(3)))
                status = "마감" if now > dl else "현재 신청 가능"
            except Exception:
                status = "현재 신청 가능"
        else:
            status = "현재 신청 가능"

    # ── spots
    spots: int | None = None
    sm = re.search(r"(\d+)\s*명", content_text)
    if sm:
        spots = int(sm.group(1))

    # ── support
    support: list[str] = []
    if cost and re.search(r"무료", cost):
        support.append("무료")
    elif cost:
        support.append(f"참여비용: {cost}")
    if place:
        support.append(f"장소: {place[:50]}")
    support = support[:5]

    # ── cover: 힌트 or 본문 첫 이미지
    cover = hint.get("cover", "")
    if not cover and content:
        img = content.select_one("img[src*='/pds/']")
        if img:
            src = img.get("src", "")
            cover = urljoin(BASE_URL, src) if src.startswith("/") else src

    # ── tag: 카테고리
    tag = hint.get("category", "")

    # ── summary: 본문 도입 문장 (▶ 이전 텍스트)
    summary = ""
    pre_arrow = re.split(r"▶", content_text)[0] if "▶" in content_text else content_text
    for sent in re.split(r"[\n。]", pre_arrow):
        s = clean(sent)
        if len(s) > 15 and not re.match(r"공지|센터사업|조회|등록", s):
            summary = s[:120]
            break

    # ── detail
    detail = content_text[:600]

    # ── duration
    duration = period
    m = re.search(r"(총\s*\d+회|\d+주|\d+개월|매주)", content_text)
    if m:
        duration = (duration + " / " + m.group()).strip(" /")

    # ── bullets: ▶ 항목 전체
    bullets: list[str] = []
    for mb in re.finditer(r"▶\s*([^\n▶☎]{5,80})", content_text):
        b = clean(mb.group(1))
        if b and b not in bullets:
            bullets.append(b)
    bullets = bullets[:6]

    return {
        "title":     title,
        "org":       org,
        "region":    region,
        "mode":      mode,
        "motive":    motive,
        "status":    status,
        "period":    period,
        "support":   support,
        "deadline":  deadline,
        "duration":  duration,
        "spots":     spots,
        "spotsLeft": None,
        "cover":     cover,
        "tag":       tag,
        "summary":   summary,
        "detail":    detail,
        "bullets":   bullets,
        "contact":   contact,
        "pub_date":  pub_date,
        "url":       url,
    }


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

COLUMNS = [
    "title", "org", "region", "mode", "motive", "status",
    "period", "support", "deadline", "duration", "spots",
    "spotsLeft", "cover", "tag", "summary", "detail", "bullets",
    "contact", "pub_date", "url",
]
COL_WIDTHS = {
    "title": 40, "org": 24, "region": 10, "mode": 10, "motive": 35,
    "status": 14, "period": 28, "support": 28, "deadline": 20,
    "duration": 28, "spots": 8, "spotsLeft": 10, "cover": 28,
    "tag": 14, "summary": 42, "detail": 50, "bullets": 40,
    "contact": 22, "pub_date": 14, "url": 32,
}
CENTER_COLS = {"region", "mode", "status", "spots", "spotsLeft", "pub_date"}


def save_excel(results: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공지사항"

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    side   = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    odd    = PatternFill("solid", start_color="EBF3FA")
    even   = PatternFill("solid", start_color="FFFFFF")

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 25

    for ri, ev in enumerate(results, 2):
        fill = odd if ri % 2 == 1 else even
        for ci, col in enumerate(COLUMNS, 1):
            val = ev.get(col)
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in CENTER_COLS else left
            if col == "url" and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
    print(f"  ✅ 엑셀 저장 완료 → {path}")


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="대구광역시 청년센터 공지사항 스크래퍼")
    parser.add_argument("--pages", type=int, default=133, help="최대 페이지 수 (기본: 133)")
    args = parser.parse_args()

    print("=" * 60)
    print("  대구광역시 청년센터 공지사항 스크래퍼")
    print(f"  최대 {args.pages}페이지 순회")
    print("=" * 60)

    print("\n[1/3] 목록에서 게시물 링크 수집 중...")
    hints = get_post_links(max_pages=args.pages)
    print(f"  → {len(hints)}개 게시물 발견")

    if not hints:
        print("  ⚠️  게시물을 찾지 못했습니다.")
        return

    print("\n[2/3] 상세 정보 추출 중...")
    results = []
    for i, hint in enumerate(hints, 1):
        print(f"  ({i:04d}/{len(hints)}) {hint['title'][:50]}...")
        detail = parse_detail(hint)
        if detail:
            results.append(detail)
        time.sleep(random.uniform(0.4, 0.8))

    print("\n[3/3] 저장 중...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"dgyouth_events_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 저장 완료 → {json_path}")

    save_excel(results, f"dgyouth_events_{ts}.xlsx")

    print("\n" + "=" * 60)
    print(json.dumps(results[:3], ensure_ascii=False, indent=2))  # 첫 3개만 미리보기
    print(f"  ... 외 {len(results)-3}개")
    print("=" * 60)
    print(f"\n✅ 완료! 총 {len(results)}개 게시물")


if __name__ == "__main__":
    main()
