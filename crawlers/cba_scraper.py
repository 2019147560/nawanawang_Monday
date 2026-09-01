"""
충청북도기업진흥원(cba.ne.kr) 사업공고/신청 스크래퍼
- 은둔고립청년 / 청년일자리 / 청년모임 관련 게시물만 수집
- 사이트 내 키워드 검색 기능 활용
사용법: python cba_scraper.py
결과물: cba_events_YYYYMMDD_HHMMSS.json / .xlsx
"""

import re
import time
import json
import random
from datetime import datetime
from urllib.parse import urljoin, urlencode

import requests
from bs4 import BeautifulSoup

BASE_URL = "https://www.cba.ne.kr"
# 사업공고 게시판 (menukey=172), 키워드 검색
SEARCH_URL = "https://www.cba.ne.kr/home/sub.php"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": BASE_URL,
}

ORG_NAME = "충청북도기업진흥원"

# ── 관련 키워드 목록 ──────────────────────────────────────────────────────────
# 사이트 검색 시 사용할 키워드 (각각 별도 검색)
SEARCH_KEYWORDS = [
    "은둔",
    "고립",
    "청년일자리",
    "청년 모임",
    "청년취업",
    "청년지원",
]

# 수집된 게시물 제목/본문에서 관련성 판단할 키워드
RELEVANT_KEYWORDS = [
    "은둔", "고립", "히키코모리", "사회적고립",
    "청년일자리", "청년 일자리", "일자리지원", "취업지원", "취업연계",
    "청년모임", "청년 모임", "자조모임", "커뮤니티", "동아리",
    "청년취업", "청년고용", "청년채용",
    "니트", "NEET",
]


# ── 유틸 ─────────────────────────────────────────────────────────────────────

def fetch(url: str, params: dict = None) -> BeautifulSoup | None:
    try:
        r = requests.get(url, params=params, headers=HEADERS, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [오류] {url} → {e}")
        return None


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def is_relevant(text: str) -> bool:
    """제목/본문에 관련 키워드가 포함되어 있는지 판단"""
    text_lower = text.lower()
    return any(kw.lower() in text_lower for kw in RELEVANT_KEYWORDS)


# ── 1단계: 키워드 검색으로 관련 게시물 링크 수집 ─────────────────────────────

def get_post_links() -> list[dict]:
    """
    cba.ne.kr 사업공고 게시판 검색:
    ?menukey=172&mod=list&search=1&kwd=키워드&page=N
    search 파라미터: 1=제목, 9=제목+내용
    """
    links = []
    seen  = set()

    for kwd in SEARCH_KEYWORDS:
        print(f"  키워드 '{kwd}' 검색 중...")

        for page in range(1, 20):
            params = {
                "menukey": "172",
                "mod":     "list",
                "search":  "9",     # 제목+내용 검색
                "kwd":     kwd,
                "page":    page,
            }
            soup = fetch(SEARCH_URL, params=params)
            if not soup:
                break

            # 게시판 목록 행: tr 또는 li 안의 링크
            rows = (
                soup.select("table.board_list tbody tr") or
                soup.select("ul.board_list li") or
                soup.select("tr.list_tr") or
                soup.select("div.board_list > ul > li")
            )

            # fallback: mod=view 링크 전체 탐색
            if not rows:
                anchors = soup.select("a[href*='mod=view']")
                rows = [a.find_parent("tr") or a.find_parent("li") or a
                        for a in anchors]

            found_any = False
            for row in rows:
                # 링크 추출
                a = None
                if hasattr(row, "select_one"):
                    a = row.select_one("a[href*='mod=view']")
                elif row.name == "a":
                    a = row

                if not a:
                    continue

                href = a.get("href", "")
                full_url = urljoin(BASE_URL, href)
                if full_url in seen:
                    continue

                title = clean(a.get_text())
                if not title or len(title) < 3:
                    if hasattr(row, "get_text"):
                        title = clean(row.get_text())[:80]

                # 제목 기준 관련성 1차 필터
                if not is_relevant(title) and not is_relevant(kwd):
                    # 키워드 자체가 관련 키워드면 통과
                    pass

                seen.add(full_url)
                links.append({
                    "url":     full_url,
                    "title":   title,
                    "keyword": kwd,
                    "cover":   "",
                })
                found_any = True

            # 검색 결과 없거나 마지막 페이지
            if not found_any:
                break

            # 페이지네이션 확인 (다음 페이지 없으면 중단)
            next_btn = soup.select_one("a[href*='page={}']".format(page + 1))
            if not next_btn:
                break

            time.sleep(random.uniform(0.5, 1.0))

        time.sleep(random.uniform(0.3, 0.7))

    print(f"  → 총 {len(links)}개 후보 게시물 수집")
    return links


# ── 2단계: 상세 페이지 파싱 + 관련성 2차 필터 ───────────────────────────────

def parse_detail(hint: dict) -> dict | None:
    url  = hint["url"]
    soup = fetch(url)
    if not soup:
        return None

    # ── title
    title_tag = (
        soup.select_one(".board_view_title") or
        soup.select_one(".view_title") or
        soup.select_one("h3.title") or
        soup.select_one("h2") or
        soup.select_one("td.title")
    )
    title = clean(title_tag.get_text()) if title_tag else hint.get("title", "")

    # ── 본문
    content = (
        soup.select_one(".board_view_content") or
        soup.select_one(".view_content") or
        soup.select_one(".content") or
        soup.select_one("td.content")
    )
    content_text = clean(content.get_text(" ")) if content else ""

    # ── 관련성 2차 필터: 제목+본문 모두 관련 키워드 없으면 skip
    if not is_relevant(title) and not is_relevant(content_text[:500]):
        return None

    # ── 게시일
    pub_date = ""
    date_patterns = [
        r"등록일[:\s]*(20\d{2}[.\-/]\d{2}[.\-/]\d{2})",
        r"작성일[:\s]*(20\d{2}[.\-/]\d{2}[.\-/]\d{2})",
        r"(20\d{2})[.\-/](\d{2})[.\-/](\d{2})",
    ]
    for pat in date_patterns:
        m = re.search(pat, soup.get_text())
        if m:
            pub_date = m.group(1) if len(m.groups()) == 1 else f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            break

    # ── org
    org = ORG_NAME
    m = re.search(r"(주관|주최|담당기관|운영기관)[:\s]*([^\n,/]{2,30})", content_text)
    if m:
        org = clean(m.group(2))

    # ── region: 충북 기본값
    region = "충북"
    region_list = ["전국", "청주", "충주", "제천", "보은", "옥천", "영동",
                   "증평", "진천", "괴산", "음성", "단양"]
    for r in region_list:
        if r in content_text:
            region = r
            break

    # ── mode
    if re.search(r"온·오프|온\s*오프|혼합", content_text):
        mode = "혼합"
    elif re.search(r"온라인|비대면|zoom|화상", content_text, re.I):
        mode = "온라인"
    elif re.search(r"오프라인|현장|대면|방문", content_text):
        mode = "오프라인"
    else:
        mode = ""

    # ── motive: 지원대상
    motive: list[str] = []
    m = re.search(
        r"(지원\s*대상|신청\s*대상|모집\s*대상|참여\s*대상)[^\n:□]*[:\s□]+([\s\S]{5,300}?)(?=\n\n|□|○|▶|\Z)",
        content_text
    )
    if m:
        block = m.group(2)
        items = re.split(r"[\n,、◦]", block)
        motive = [clean(i) for i in items if 3 < len(clean(i)) < 100][:5]

    # ── deadline: 신청기간
    deadline = ""
    m = re.search(
        r"(신청\s*기간|모집\s*기간|접수\s*기간)[^\n:□]*[:\s□]+([\d.\-년월일\s~]+)",
        content_text
    )
    if m:
        deadline = clean(m.group(2))

    # ── status
    if re.search(r"마감|종료|접수종료|모집완료", title + content_text[:300]):
        status = "마감"
    elif re.search(r"예정|추후공지|미정", content_text[:300]):
        status = "모집 예정"
    else:
        now = datetime.now()
        date_m = re.search(r"(20\d{2})[.\-년]\s*(\d{1,2})[.\-월]\s*(\d{1,2})", deadline)
        if date_m:
            try:
                dl = datetime(int(date_m.group(1)), int(date_m.group(2)), int(date_m.group(3)))
                status = "마감" if now > dl else "현재 신청 가능"
            except Exception:
                status = "현재 신청 가능"
        else:
            status = "현재 신청 가능"

    # ── period
    period = ""
    m = re.search(
        r"(사업\s*기간|운영\s*기간|지원\s*기간|프로그램\s*기간)[^\n:□]*[:\s□]+([\d.\-년월일\s~()가-힣]+)",
        content_text
    )
    if m:
        period = clean(m.group(2))[:80]

    # ── support: 지원내용
    support: list[str] = []
    m = re.search(
        r"(지원\s*내용|지원\s*사항|혜택|주요\s*내용)[^\n:□]*[:\s□]+([\s\S]{10,400}?)(?=\n\n|□|○|\Z)",
        content_text
    )
    if m:
        block = m.group(2)
        items = re.split(r"[\n,◦]", block)
        support = [clean(i) for i in items if 4 < len(clean(i)) < 100][:6]

    # ── duration
    duration = period
    m = re.search(r"(총\s*\d+회|\d+개월|\d+주|매주)", content_text)
    if m:
        duration = (duration + " / " + m.group()).strip(" /")

    # ── spots
    spots: int | None = None
    m = re.search(r"(\d+)\s*명\s*(이내|모집|선발|내외|정원)", content_text)
    if m:
        spots = int(m.group(1))

    # ── cover: 첨부 이미지
    cover = ""
    if content:
        img = content.select_one("img")
        if img:
            src = img.get("src", "")
            cover = urljoin(BASE_URL, src) if src.startswith("/") else src

    # ── tag: 카테고리 (청년지원, 일자리지원 등)
    tag = ""
    cat_tag = soup.select_one(".category, .cate, .board_cate")
    if cat_tag:
        tag = clean(cat_tag.get_text())
    if not tag:
        # 관련 키워드 중 매칭된 것
        for kw in RELEVANT_KEYWORDS:
            if kw.lower() in title.lower():
                tag = kw
                break

    # ── summary
    summary = ""
    for sent in re.split(r"[\n。]", content_text):
        s = clean(sent)
        if len(s) > 20:
            summary = s[:120]
            break

    # ── detail
    detail = content_text[:600]

    # ── bullets: □ / ○ / ▶ 항목
    bullets: list[str] = []
    for mb in re.finditer(r"[□○▶◇•]\s*([^\n□○▶◇•]{5,80})", content_text):
        b = clean(mb.group(1))
        if b and b not in bullets:
            bullets.append(b)
    if not bullets:
        for mb in re.finditer(r"\d+[..]\s*([^\n]{5,80})", content_text):
            b = clean(mb.group(1))
            if b and b not in bullets:
                bullets.append(b)
    bullets = bullets[:6]

    return {
        "title":     title,
        "org":       org,
        "region":    region,
        "mode":      mode,
        "motive":    motive,
        "status":    status,
        "period":    period,
        "support":   support,
        "deadline":  deadline,
        "duration":  duration,
        "spots":     spots,
        "spotsLeft": None,
        "cover":     cover,
        "tag":       tag,
        "summary":   summary,
        "detail":    detail,
        "bullets":   bullets,
        "pub_date":  pub_date,
        "url":       url,
    }


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

COLUMNS = [
    "title", "org", "region", "mode", "motive", "status",
    "period", "support", "deadline", "duration", "spots",
    "spotsLeft", "cover", "tag", "summary", "detail", "bullets",
    "pub_date", "url",
]
COL_WIDTHS = {
    "title": 38, "org": 20, "region": 10, "mode": 12, "motive": 35,
    "status": 14, "period": 28, "support": 35, "deadline": 20,
    "duration": 28, "spots": 10, "spotsLeft": 10, "cover": 30,
    "tag": 14, "summary": 42, "detail": 52, "bullets": 42,
    "pub_date": 14, "url": 32,
}
CENTER_COLS = {"region", "mode", "status", "spots", "spotsLeft", "tag", "pub_date"}


def save_excel(results: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "사업공고"

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    side   = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    odd    = PatternFill("solid", start_color="EBF3FA")
    even   = PatternFill("solid", start_color="FFFFFF")

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 25

    for ri, ev in enumerate(results, 2):
        fill = odd if ri % 2 == 1 else even
        for ci, col in enumerate(COLUMNS, 1):
            val = ev.get(col)
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in CENTER_COLS else left
            if col == "url" and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
    print(f"  ✅ 엑셀 저장 완료 → {path}")


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  충청북도기업진흥원 사업공고 관련 게시물 스크래퍼")
    print(f"  검색 키워드: {', '.join(SEARCH_KEYWORDS)}")
    print("=" * 60)

    print("\n[1/3] 키워드 검색으로 관련 게시물 링크 수집 중...")
    hints = get_post_links()

    print("\n[2/3] 상세 정보 추출 + 관련성 2차 필터링 중...")
    results = []
    skipped = 0
    for i, hint in enumerate(hints, 1):
        print(f"  ({i:02d}/{len(hints)}) {hint['title'][:50]}...")
        detail = parse_detail(hint)
        if detail:
            results.append(detail)
        else:
            skipped += 1
            print(f"    → 관련성 없음, skip")
        time.sleep(random.uniform(0.5, 1.0))

    print(f"\n  → 최종 {len(results)}개 수집 / {skipped}개 필터링됨")

    if not results:
        print("  ⚠️  관련 게시물을 찾지 못했습니다.")
        return

    print("\n[3/3] 저장 중...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"cba_events_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 저장 완료 → {json_path}")

    save_excel(results, f"cba_events_{ts}.xlsx")

    print("\n" + "=" * 60)
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print("=" * 60)
    print(f"\n✅ 완료! 총 {len(results)}개 관련 게시물")


if __name__ == "__main__":
    main()
