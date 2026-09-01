"""
울산중구종합사회복지관(jsbwelfare.or.kr) 공지사항 스크래퍼
- 그누보드(board.php?bo_table=notice) 표준 구조 파싱
- 전체 게시물 수집
사용법: python jsbwelfare_scraper.py
결과물: jsbwelfare_events_YYYYMMDD_HHMMSS.json / .xlsx
"""

import re
import time
import json
import random
import warnings
from datetime import datetime
from urllib.parse import urljoin, urlencode

import requests
from bs4 import BeautifulSoup

# SSL 경고 억제
warnings.filterwarnings("ignore")

BASE_URL = "http://www.jsbwelfare.or.kr"
LIST_URL = "http://www.jsbwelfare.or.kr/bbs/board.php"

LIST_PARAMS = {
    "bo_table": "notice",
    "sst":      "wr_datetime",
    "sod":      "desc",
    "sop":      "and",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": BASE_URL,
}

ORG_NAME = "울산중구종합사회복지관"


# ── 유틸 ─────────────────────────────────────────────────────────────────────

def fetch(url: str, params: dict = None) -> BeautifulSoup | None:
    try:
        r = requests.get(
            url, params=params, headers=HEADERS,
            timeout=15, verify=False
        )
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [오류] {url} → {e}")
        return None


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


# ── 1단계: 목록 페이지에서 wr_id 수집 ────────────────────────────────────────

def get_post_links() -> list[dict]:
    """
    그누보드 목록 구조:
    - .bo_tbl tbody tr > td.td_subject > a[href*='wr_id=']
    - 또는 카드형: .gall_tit > a[href*='wr_id=']
    페이지네이션: &page=N
    """
    links = []
    seen  = set()

    for page in range(1, 200):  # 충분히 크게
        params = {**LIST_PARAMS, "page": page}
        print(f"  목록 페이지 {page} 파싱 중...")
        soup = fetch(LIST_URL, params=params)
        if not soup:
            break

        # 그누보드 표준: td.td_subject a 또는 갤러리형: .gall_tit a
        anchors = (
            soup.select("td.td_subject a[href*='wr_id=']") or
            soup.select(".gall_tit a[href*='wr_id=']") or
            soup.select(".bo_tit a[href*='wr_id=']") or
            soup.select("a[href*='bo_table=notice'][href*='wr_id=']")
        )

        found_any = False
        for a in anchors:
            href = a.get("href", "")
            m = re.search(r"wr_id=(\d+)", href)
            if not m:
                continue
            wr_id = m.group(1)
            if wr_id in seen:
                continue

            full_url = urljoin(BASE_URL, href)
            title = clean(a.get_text())

            # 행에서 추가 정보
            row = a.find_parent("tr") or a.find_parent("li") or a.find_parent("div")
            pub_date = ""
            if row:
                # 날짜 셀
                date_td = row.select_one("td.td_date, .td_datetime, time")
                if date_td:
                    pub_date = clean(date_td.get_text())
                else:
                    dm = re.search(r"(20\d{2}[.\-]\d{2}[.\-]\d{2})", clean(row.get_text()))
                    pub_date = dm.group(1) if dm else ""

            # 썸네일 (갤러리형)
            img = a.select_one("img") or (row.select_one("img") if row else None)
            cover = ""
            if img:
                src = img.get("src", "")
                cover = urljoin(BASE_URL, src) if src else ""

            seen.add(wr_id)
            links.append({
                "url":      full_url,
                "wr_id":    wr_id,
                "title":    title,
                "pub_date": pub_date,
                "cover":    cover,
            })
            found_any = True

        if not found_any:
            print(f"  → 페이지 {page}에 게시물 없음, 중단")
            break

        # 다음 페이지 존재 여부
        next_link = soup.select_one(f"a[href*='page={page+1}']")
        if not next_link:
            print(f"  → 마지막 페이지 ({page})")
            break

        time.sleep(random.uniform(0.5, 1.0))

    return links


# ── 2단계: 상세 페이지 파싱 ──────────────────────────────────────────────────

def parse_detail(hint: dict) -> dict:
    url  = hint["url"]
    soup = fetch(url)
    if not soup:
        return {}

    # ── title: 그누보드 표준 .bo_v_tit 또는 h2
    title_tag = (
        soup.select_one(".bo_v_tit") or
        soup.select_one("#bo_v_title") or
        soup.select_one("h2.title") or
        soup.select_one("h1")
    )
    title = clean(title_tag.get_text()) if title_tag else hint.get("title", "")
    # [마감] [대기신청가능] 등 앞 배지 보존, 하지만 태그 정리
    title = re.sub(r"\s+", " ", title).strip()

    # ── 등록일: 그누보드 .bo_v_info
    pub_date = hint.get("pub_date", "")
    info = soup.select_one(".bo_v_info, #bo_v_info, .view-info")
    if info:
        dm = re.search(r"(20\d{2}[.\-]\d{2}[.\-]\d{2})", clean(info.get_text()))
        if dm:
            pub_date = dm.group(1).replace(".", "-")

    # ── 본문: 그누보드 표준 #bo_v_con
    content = (
        soup.select_one("#bo_v_con") or
        soup.select_one(".bo_v_con") or
        soup.select_one(".view_content") or
        soup.select_one("section#content")
    )
    content_text = clean(content.get_text(" ")) if content else ""

    # ── 상태: 제목에서 [마감] [대기] 등 배지 확인
    if re.search(r"\[마감\]|\[모집마감\]|마감|접수종료|모집완료", title):
        status = "마감"
    elif re.search(r"\[대기\]|\[대기신청\]", title):
        status = "마감"  # 대기 = 사실상 마감
    elif re.search(r"예정|추후공지", content_text[:300]):
        status = "모집 예정"
    else:
        # 마감일 기준 판단
        now = datetime.now()
        dl_m = re.search(
            r"(마감|신청기간|모집기간)[^\d]*(20\d{2})[.\-년]\s*(\d{1,2})[.\-월]\s*(\d{1,2})",
            content_text
        )
        if dl_m:
            try:
                dl = datetime(int(dl_m.group(2)), int(dl_m.group(3)), int(dl_m.group(4)))
                status = "마감" if now > dl else "현재 신청 가능"
            except Exception:
                status = "현재 신청 가능"
        else:
            status = "현재 신청 가능"

    # ── org
    org = ORG_NAME
    m = re.search(r"(주관|주최|운영기관)[:\s]*([^\n,/]{2,30})", content_text)
    if m:
        val = clean(m.group(2))
        if val and val != ORG_NAME:
            org = f"{ORG_NAME} / {val}"

    # ── region
    region = "울산"
    region_list = ["중구", "남구", "동구", "북구", "울주군"]
    for r in region_list:
        if r in content_text:
            region = f"울산 {r}"
            break

    # ── mode
    if re.search(r"온·오프|온\s*오프|혼합", content_text):
        mode = "혼합"
    elif re.search(r"온라인|비대면|zoom|유튜브|카카오", content_text, re.I):
        mode = "온라인"
    elif re.search(r"오프라인|현장|대면|복지관|센터|현장", content_text):
        mode = "오프라인"
    else:
        mode = ""

    # ── motive: 대상
    motive: list[str] = []
    m = re.search(
        r"(대\s*상|신청자격|참여자격|모집대상)[^\n:◎○▶•]*[:\s◎○▶•]+([\s\S]{5,300}?)"
        r"(?=\n\n|◎|○|▶|•|\*|일\s*시|기\s*간|장\s*소|\Z)",
        content_text
    )
    if m:
        items = re.split(r"[\n,、◦·]", m.group(2))
        motive = [clean(i) for i in items if 3 < len(clean(i)) < 100][:5]

    # ── deadline
    deadline = ""
    m = re.search(
        r"(신청기간|모집기간|접수기간|마감일)[^\n:◎○▶•]*[:\s◎○▶•]+([\d.\-년월일\s~()가-힣]+)",
        content_text
    )
    if m:
        raw = clean(m.group(2))
        deadline = raw.split("~")[-1].strip() if "~" in raw else raw

    # ── period: 일시/운영기간
    period = ""
    m = re.search(
        r"(일\s*시|운영기간|프로그램기간|활동기간|기\s*간)[^\n:◎○▶•]*[:\s◎○▶•]+"
        r"([\d.\-년월일()\s~가-힣요]+)",
        content_text
    )
    if m:
        period = clean(m.group(2))[:80]

    # ── support
    support: list[str] = []
    if re.search(r"무료|참가비\s*없음|수강료\s*없음", content_text):
        support.append("무료")
    m = re.search(
        r"(지원내용|지원사항|혜택|제공)[^\n:◎○▶•]*[:\s◎○▶•]+([\s\S]{5,300}?)"
        r"(?=\n\n|◎|○|▶|\Z)",
        content_text
    )
    if m:
        items = re.split(r"[\n,]", m.group(2))
        support += [clean(i) for i in items if 4 < len(clean(i)) < 80]
    support = list(dict.fromkeys(support))[:5]

    # ── duration
    duration = period
    m = re.search(r"(총\s*\d+회|\d+박\s*\d+일|\d+주|\d+개월|매주)", content_text)
    if m:
        duration = (duration + " / " + m.group()).strip(" /")

    # ── spots
    spots: int | None = None
    m = re.search(r"(\d+)\s*명\s*(이내|모집|선발|내외|정원|한정)", content_text)
    if m:
        spots = int(m.group(1))

    # ── cover
    cover = hint.get("cover", "")
    if not cover and content:
        img = content.select_one("img[src]")
        if img:
            src = img.get("src", "")
            cover = urljoin(BASE_URL, src) if src.startswith("/") else src

    # ── tag: 제목 앞 [] 배지
    tag = ""
    tm = re.search(r"^\[([^\]]{1,20})\]", title)
    if tm:
        tag = tm.group(1)

    # ── summary
    summary = ""
    for sent in re.split(r"[\n。]", content_text):
        s = clean(sent)
        if len(s) > 15 and not re.match(r"등록|조회|댓글|첨부", s):
            summary = s[:120]
            break

    # ── detail
    detail = content_text[:600]

    # ── bullets: ◎ / ○ / ▶ / • / - 항목
    bullets: list[str] = []
    for mb in re.finditer(r"[◎○▶•◆■]\s*([^\n◎○▶•◆■]{5,80})", content_text):
        b = clean(mb.group(1))
        if b and b not in bullets:
            bullets.append(b)
    if not bullets:
        for mb in re.finditer(r"[-–]\s+([^\n\-]{5,80})", content_text):
            b = clean(mb.group(1))
            if b and b not in bullets:
                bullets.append(b)
    bullets = bullets[:6]

    # ── contact
    contact = ""
    m = re.search(r"(문의|연락|☎|전화)[^\n:]*[:\s]*([\d\-()]+)", content_text)
    if m:
        contact = clean(m.group(2))

    return {
        "title":     title,
        "org":       org,
        "region":    region,
        "mode":      mode,
        "motive":    motive,
        "status":    status,
        "period":    period,
        "support":   support,
        "deadline":  deadline,
        "duration":  duration,
        "spots":     spots,
        "spotsLeft": None,
        "cover":     cover,
        "tag":       tag,
        "summary":   summary,
        "detail":    detail,
        "bullets":   bullets,
        "contact":   contact,
        "pub_date":  pub_date,
        "url":       url,
    }


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

COLUMNS = [
    "title", "org", "region", "mode", "motive", "status",
    "period", "support", "deadline", "duration", "spots",
    "spotsLeft", "cover", "tag", "summary", "detail", "bullets",
    "contact", "pub_date", "url",
]
COL_WIDTHS = {
    "title": 40, "org": 24, "region": 10, "mode": 10, "motive": 35,
    "status": 14, "period": 28, "support": 28, "deadline": 20,
    "duration": 28, "spots": 8, "spotsLeft": 10, "cover": 28,
    "tag": 16, "summary": 42, "detail": 50, "bullets": 38,
    "contact": 18, "pub_date": 14, "url": 32,
}
CENTER_COLS = {"region", "mode", "status", "spots", "spotsLeft", "pub_date"}


def save_excel(results: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공지사항"

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    side   = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    odd    = PatternFill("solid", start_color="EBF3FA")
    even   = PatternFill("solid", start_color="FFFFFF")

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 25

    for ri, ev in enumerate(results, 2):
        fill = odd if ri % 2 == 1 else even
        for ci, col in enumerate(COLUMNS, 1):
            val = ev.get(col)
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in CENTER_COLS else left
            if col == "url" and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
    print(f"  ✅ 엑셀 저장 완료 → {path}")


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  울산중구종합사회복지관 공지사항 스크래퍼")
    print("=" * 60)

    print("\n[1/3] 목록에서 게시물 링크 수집 중...")
    hints = get_post_links()
    print(f"  → {len(hints)}개 게시물 발견")

    if not hints:
        print("  ⚠️  게시물을 찾지 못했습니다.")
        return

    print("\n[2/3] 상세 정보 추출 중...")
    results = []
    for i, hint in enumerate(hints, 1):
        print(f"  ({i:04d}/{len(hints)}) {hint['title'][:50]}...")
        detail = parse_detail(hint)
        if detail:
            results.append(detail)
        time.sleep(random.uniform(0.4, 0.9))

    print("\n[3/3] 저장 중...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"jsbwelfare_events_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 저장 완료 → {json_path}")

    save_excel(results, f"jsbwelfare_events_{ts}.xlsx")

    print("\n" + "=" * 60)
    print(json.dumps(results[:3], ensure_ascii=False, indent=2))
    if len(results) > 3:
        print(f"  ... 외 {len(results)-3}개")
    print("=" * 60)
    print(f"\n✅ 완료! 총 {len(results)}개 게시물")


if __name__ == "__main__":
    main()
