"""
행복공장(happitory.org) 공지사항 스크래퍼
- XpressEngine 기반 사이트 구조 파싱
- 날짜 필터 없이 전체 수집
사용법: python happitory_scraper.py
결과물: happitory_events_YYYYMMDD_HHMMSS.json / .xlsx
"""

import re
import time
import json
import random
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

BASE_URL  = "https://www.happitory.org"
LIST_URL  = "https://www.happitory.org/comm_notice"
PAGE_URL  = "https://www.happitory.org/index.php?mid=comm_notice&page={page}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": BASE_URL,
}

ORG_NAME    = "(사)행복공장"
ORG_CONTACT = "02-6084-1016 / hf1016@daum.net"
ORG_REGION  = "서울"


# ── 유틸 ─────────────────────────────────────────────────────────────────────

def fetch(url: str) -> BeautifulSoup | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [오류] {url} → {e}")
        return None


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


# ── 1단계: 목록 페이지에서 게시물 링크 수집 ─────────────────────────────────

def get_post_links() -> list[dict]:
    """
    XpressEngine 목록 구조:
    - 공지글: li.notice > a[href*='/comm_notice/']
    - 일반글: 카드형 a[href*='/comm_notice/숫자']
    페이지네이션: ?mid=comm_notice&page=N (총 36페이지)
    """
    links = []
    seen  = set()

    # 1페이지는 기본 URL, 이후는 page 파라미터
    urls = [LIST_URL] + [PAGE_URL.format(page=p) for p in range(2, 37)]

    for page_num, url in enumerate(urls, 1):
        print(f"  목록 페이지 {page_num} 파싱 중...")
        soup = fetch(url)
        if not soup:
            break

        # comm_notice/숫자 패턴 링크만 수집
        found_any = False
        for a in soup.select("a[href*='/comm_notice/']"):
            href = a.get("href", "")
            # /comm_notice/숫자 형태만
            if not re.search(r"/comm_notice/\d+", href):
                continue
            full_url = urljoin(BASE_URL, href)
            if full_url in seen:
                continue

            title = clean(a.get_text())
            if not title or len(title) < 2:
                # 이미지만 있는 링크 → 상위 요소에서 제목 추출
                parent = a.find_parent(["li", "div", "article"])
                if parent:
                    h = parent.select_one("h4, h5, strong, p")
                    title = clean(h.get_text()) if h else ""

            # 썸네일
            img = a.select_one("img[src*='thumbnails'], img[src*='attach']")
            cover = img.get("src", "") if img else ""

            seen.add(full_url)
            links.append({
                "url":   full_url,
                "title": title,
                "cover": cover,
            })
            found_any = True

        if not found_any:
            print(f"  → 페이지 {page_num}에 게시물 없음, 중단")
            break

        time.sleep(random.uniform(0.4, 0.8))

    return links


# ── 2단계: 상세 페이지 파싱 ──────────────────────────────────────────────────

def parse_detail(hint: dict) -> dict:
    url  = hint["url"]
    soup = fetch(url)
    if not soup:
        return {}

    # ── title: h2.xe_content 또는 페이지 h2
    h2 = soup.select_one("h2")
    title = clean(h2.get_text()) if h2 else hint["title"]
    # " - 공지사항 - 행복공장" 접미사 제거
    title = re.sub(r"\s*[-–]\s*공지사항.*$", "", title).strip()

    # ── 게시일: meta article:published_time
    pub_meta = soup.find("meta", attrs={"name": "article:published_time"}) or \
               soup.find("meta", property="article:published_time")
    pub_date = ""
    if pub_meta:
        pub_date = pub_meta.get("content", "")[:10]  # YYYY-MM-DD
    # fallback: 본문 날짜 텍스트
    if not pub_date:
        dm = re.search(r"(20\d{2})\.(0?\d|1[0-2])\.(0?\d|[12]\d|3[01])", soup.get_text())
        if dm:
            pub_date = dm.group(0)

    # ── 본문 영역
    content = (
        soup.select_one(".xe_content") or
        soup.select_one(".board-read-body") or
        soup.select_one("article") or
        soup.select_one(".rd_body")
    )
    content_text = clean(content.get_text(" ")) if content else ""

    # ── org
    org = ORG_NAME
    m = re.search(r"주\s*최[:\s]*([^\n/,]{2,30})", content_text)
    if m:
        org = clean(m.group(1))

    # ── region
    region = ORG_REGION
    region_list = ["전국", "부산", "대구", "인천", "광주", "대전", "울산",
                   "세종", "경기", "강원", "충북", "충남", "전북", "전남",
                   "경북", "경남", "제주", "홍천"]
    for r in region_list:
        if r in content_text:
            region = r
            break

    # ── mode
    if re.search(r"온·오프|온\s*오프|혼합", content_text):
        mode = "혼합"
    elif re.search(r"온라인|비대면|zoom|유튜브|화상", content_text, re.I):
        mode = "온라인"
    elif re.search(r"오프라인|현장|대면|캠프|수련원|극장|홀", content_text):
        mode = "오프라인"
    else:
        mode = ""

    # ── motive: 대상 항목
    motive: list[str] = []
    m = re.search(r"대\s*상[:\s*◇○▶•]*([^\n◇○▶]{5,200})", content_text)
    if m:
        block = m.group(1)
        items = re.split(r"[,、\n]", block)
        motive = [clean(i) for i in items if 3 < len(clean(i)) < 80][:5]

    # ── deadline: 신청마감일
    deadline = ""
    m = re.search(r"(신청\s*마감|모집\s*마감|접수\s*마감)[^\d]*([\d.년월일\s~\-]+)", content_text)
    if m:
        deadline = clean(m.group(2))

    # ── status: [신청마감] 태그 우선
    if re.search(r"\[신청마감\]|\[마감\]|신청마감|모집완료|접수종료", title + content_text[:200]):
        status = "마감"
    elif re.search(r"예정|추후공지|미정", content_text[:200]):
        status = "모집 예정"
    else:
        now = datetime.now()
        date_m = re.search(r"(20\d{2})[.\s]\s*(\d{1,2})[.\s]\s*(\d{1,2})", deadline)
        if date_m:
            try:
                dl = datetime(int(date_m.group(1)), int(date_m.group(2)), int(date_m.group(3)))
                status = "마감" if now > dl else "현재 신청 가능"
            except Exception:
                status = "현재 신청 가능"
        else:
            status = "현재 신청 가능"

    # ── period: ◇일시 / * 일정 섹션
    period = ""
    m = re.search(r"(일\s*시|일\s*정|기\s*간)[^\n:◇*▶]*[:\s◇*▶]+([\d.\s~\-년월일(가-힣)]+)", content_text)
    if m:
        period = clean(m.group(2))[:80]

    # ── support: 지원 내용
    support: list[str] = []
    if re.search(r"무료|수강료\s*없음|참가비\s*없음", content_text):
        support.append("무료")
    m = re.search(r"(지원|혜택|제공)[^\n:]*[:\s]+([\s\S]{10,300}?)(?=\n\n|◇|\*|후원|문의|\Z)", content_text)
    if m:
        block = m.group(2)
        items = re.split(r"[\n,]", block)
        support += [clean(i) for i in items if 4 < len(clean(i)) < 80]
    support = list(dict.fromkeys(support))[:5]

    # ── duration: 회차/기간 상세
    duration = period
    m = re.search(r"(총\s*\d+회|회\s*진행|\d+박\s*\d+일|\d+주|\d+개월)", content_text)
    if m:
        duration = (duration + " / " + m.group()).strip(" /")

    # ── spots: 모집 인원
    spots: int | None = None
    m = re.search(r"(\d+)\s*명\s*(이내|모집|선발|내외)", content_text)
    if m:
        spots = int(m.group(1))

    # ── cover: 썸네일 우선, 본문 첫 이미지 fallback
    cover = hint.get("cover", "")
    if not cover and content:
        img = content.select_one("img[src*='attach'], img[src*='files']")
        if img:
            src = img.get("src", "")
            cover = urljoin(BASE_URL, src) if src.startswith("/") else src

    # ── tag: 제목 앞 [] 태그
    tag = ""
    tm = re.search(r"^\[([^\]]{1,15})\]", title)
    if tm:
        tag = tm.group(1)

    # ── summary: 굵은 텍스트 or 첫 의미있는 문장
    summary = ""
    if content:
        bold = content.select_one("strong, b")
        if bold:
            s = clean(bold.get_text())
            if len(s) > 10:
                summary = s[:120]
    if not summary:
        for sent in re.split(r"[\n。]", content_text):
            s = clean(sent)
            if len(s) > 20 and not re.match(r"안녕하세요|행복공장입니다", s):
                summary = s[:120]
                break

    # ── detail: 본문 전체 (최대 600자)
    detail = content_text[:600]

    # ── bullets: ○ / ▶ / ◇ / • 항목
    bullets: list[str] = []
    for mb in re.finditer(r"[○▶◇•]\s*([^\n○▶◇•]{5,80})", content_text):
        b = clean(mb.group(1))
        if b and b not in bullets:
            bullets.append(b)
    if not bullets:
        for mb in re.finditer(r"-\s+([^\n\-]{5,80})", content_text):
            b = clean(mb.group(1))
            if b and b not in bullets:
                bullets.append(b)
    bullets = bullets[:6]

    return {
        "title":     title,
        "org":       org,
        "region":    region,
        "mode":      mode,
        "motive":    motive,
        "status":    status,
        "period":    period,
        "support":   support,
        "deadline":  deadline,
        "duration":  duration,
        "spots":     spots,
        "spotsLeft": None,
        "cover":     cover,
        "tag":       tag,
        "summary":   summary,
        "detail":    detail,
        "bullets":   bullets,
        "url":       url,
        "pub_date":  pub_date,
    }


# ── 엑셀 저장 ────────────────────────────────────────────────────────────────

COLUMNS = [
    "title", "org", "region", "mode", "motive", "status",
    "period", "support", "deadline", "duration", "spots",
    "spotsLeft", "cover", "tag", "summary", "detail", "bullets",
    "pub_date", "url",
]
COL_WIDTHS = {
    "title": 38, "org": 20, "region": 10, "mode": 12, "motive": 35,
    "status": 14, "period": 28, "support": 30, "deadline": 18,
    "duration": 28, "spots": 10, "spotsLeft": 10, "cover": 30,
    "tag": 14, "summary": 42, "detail": 52, "bullets": 42,
    "pub_date": 14, "url": 32,
}
CENTER_COLS = {"region", "mode", "status", "spots", "spotsLeft", "tag", "pub_date"}


def save_excel(results: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공지사항"

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    side   = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    odd    = PatternFill("solid", start_color="EBF3FA")
    even   = PatternFill("solid", start_color="FFFFFF")

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 25

    for ri, ev in enumerate(results, 2):
        fill = odd if ri % 2 == 1 else even
        for ci, col in enumerate(COLUMNS, 1):
            val = ev.get(col)
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in CENTER_COLS else left
            if col == "url" and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
    print(f"  ✅ 엑셀 저장 완료 → {path}")


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  행복공장(happitory.org) 공지사항 스크래퍼")
    print("=" * 60)

    print("\n[1/3] 목록에서 게시물 링크 수집 중...")
    hints = get_post_links()
    print(f"  → {len(hints)}개 게시물 발견")

    if not hints:
        print("  ⚠️  게시물을 찾지 못했습니다.")
        return

    print("\n[2/3] 상세 정보 추출 중...")
    results = []
    for i, hint in enumerate(hints, 1):
        print(f"  ({i:02d}/{len(hints)}) {hint['title'][:50]}...")
        detail = parse_detail(hint)
        if detail:
            results.append(detail)
        time.sleep(random.uniform(0.5, 1.0))

    print("\n[3/3] 저장 중...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"happitory_events_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 저장 완료 → {json_path}")

    save_excel(results, f"happitory_events_{ts}.xlsx")

    print("\n" + "=" * 60)
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print("=" * 60)
    print(f"\n✅ 완료! 총 {len(results)}개 게시물")


if __name__ == "__main__":
    main()
